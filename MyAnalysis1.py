# ============================================================
# FULL OLS QSPR ANALYSIS
# Linear, quadratic, and cubic one-descriptor QSPR models
# Creates: Full_Statistical_QSPR_Analysis_OLS_complete.xlsx
# Uses openpyxl only; XlsxWriter is NOT required.
# ============================================================

import sys
import warnings
from pathlib import Path

import numpy as np
import pandas as pd
import statsmodels.api as sm

from scipy.stats import pearsonr, shapiro
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_squared_error
from sklearn.model_selection import LeaveOneOut, cross_val_predict
from sklearn.preprocessing import PolynomialFeatures

from statsmodels.stats.diagnostic import het_breuschpagan
from statsmodels.stats.outliers_influence import variance_inflation_factor

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ============================================================
# 1. USER SETTINGS
# ============================================================

# Your project folder
PROJECT_FOLDER = Path(
    r"D:\Postgradute\PhD_Research New\TB_Drugs"
)

# Change this to "mydata(2).xlsx" if that is your actual file name.
INPUT_FILE_NAME = "mydata.xlsx"

# Output workbook name
OUTPUT_FILE_NAME = (
    "Full_Statistical_QSPR_Analysis_OLS_complete.xlsx"
)

# Exact input and output paths
INPUT_FILE = PROJECT_FOLDER / INPUT_FILE_NAME
OUTPUT_FILE = PROJECT_FOLDER / OUTPUT_FILE_NAME

# Name of the compound column in your Excel data file
COMPOUND_COLUMN = "Unnamed: 0"

# Physicochemical-property columns
PROPERTIES = [
    "BP",
    "MP",
    "FP",
    "EV",
    "MR",
    "PL",
    "ST",
    "MV"
]

# Linear, quadratic, and cubic models
DEGREES = [1, 2, 3]

# Overall model significance threshold
ALPHA = 0.05

# A model is considered predictively unreliable if LOOCV R2 <= 0
LOOCV_R2_THRESHOLD = 0.0

# Number of strongest Pearson correlations to keep per property
TOP_CORRELATIONS_TO_KEEP = 10


# ============================================================
# 2. GENERAL FUNCTIONS
# ============================================================

warnings.filterwarnings("ignore", category=RuntimeWarning)
warnings.filterwarnings("ignore", category=UserWarning)


def model_type_from_degree(degree):
    """Convert polynomial degree to model type name."""
    if degree == 1:
        return "Linear"
    if degree == 2:
        return "Quadratic"
    if degree == 3:
        return "Cubic"

    return f"Degree {degree}"


def format_number(value, decimals=8):
    """Format values cleanly for regression equations."""
    if pd.isna(value):
        return "NA"

    value = float(value)

    if value == 0:
        return "0"

    if abs(value) < 0.0001:
        return f"{value:.6e}"

    text = f"{value:.{decimals}f}"
    return text.rstrip("0").rstrip(".")


def make_equation(property_name, coefficients, index_name):
    """
    Produce equation such as:
    BP = 443.108 + 0.173884X + 0.000495208X^2, where X = PEB
    """
    equation = (
        f"{property_name} = "
        f"{format_number(coefficients[0])}"
    )

    for power in range(1, len(coefficients)):
        coefficient = coefficients[power]

        sign = "+" if coefficient >= 0 else "-"
        coefficient_text = format_number(abs(coefficient))

        if power == 1:
            term = f"{coefficient_text}X"
        else:
            term = f"{coefficient_text}X^{power}"

        equation += f" {sign} {term}"

    equation += f", where X = {index_name}"

    return equation


def safe_shapiro(residuals):
    """Safely calculate Shapiro-Wilk residual-normality test."""
    try:
        if len(residuals) < 3:
            return np.nan, np.nan

        statistic, p_value = shapiro(residuals)
        return statistic, p_value

    except Exception:
        return np.nan, np.nan


def safe_breusch_pagan(residuals, design_matrix):
    """Safely calculate Breusch-Pagan heteroscedasticity test."""
    try:
        lm_stat, lm_pvalue, f_stat, f_pvalue = het_breuschpagan(
            residuals,
            design_matrix
        )

        return lm_stat, lm_pvalue, f_stat, f_pvalue

    except Exception:
        return np.nan, np.nan, np.nan, np.nan


def diagnostic_status(shapiro_p, breusch_pagan_p):
    """
    Mark diagnostic status.
    'Pass' means both Shapiro-Wilk and Breusch-Pagan p-values exceed 0.05.
    """
    if pd.isna(shapiro_p) or pd.isna(breusch_pagan_p):
        return "Not assessed"

    if shapiro_p > ALPHA and breusch_pagan_p > ALPHA:
        return "Pass"

    return "Review"


def calculate_vif(design_matrix, degree):
    """
    Calculate VIF for polynomial predictor terms.
    The intercept column is excluded.
    """
    vif_rows = []

    if degree == 1:
        return [("X", 1.0)]

    for column_position in range(1, design_matrix.shape[1]):
        try:
            vif_value = variance_inflation_factor(
                design_matrix,
                column_position
            )
        except Exception:
            vif_value = np.nan

        if column_position == 1:
            term = "X"
        else:
            term = f"X^{column_position}"

        vif_rows.append((term, vif_value))

    return vif_rows


# ============================================================
# 3. OLS MODEL-FITTING FUNCTION
# ============================================================

def fit_ols_polynomial_model(
    x,
    y,
    property_name,
    index_name,
    compounds,
    degree
):
    """
    Fit one linear, quadratic, or cubic OLS model and return all outputs.
    """

    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    compounds = np.asarray(compounds, dtype=str)

    valid_mask = np.isfinite(x) & np.isfinite(y)

    x_clean = x[valid_mask]
    y_clean = y[valid_mask]
    compounds_clean = compounds[valid_mask]

    n = len(y_clean)

    # Need sufficient observations for intercept, polynomial terms,
    # and at least one residual degree of freedom.
    if n <= degree + 2:
        return None

    if np.std(x_clean) == 0:
        return None

    if np.std(y_clean) == 0:
        return None

    # Create X, X^2, and X^3
    polynomial = PolynomialFeatures(
        degree=degree,
        include_bias=False
    )

    x_poly = polynomial.fit_transform(
        x_clean.reshape(-1, 1)
    )

    # Add intercept for OLS model
    x_sm = sm.add_constant(
        x_poly,
        has_constant="add"
    )

    try:
        ols_model = sm.OLS(y_clean, x_sm).fit()
    except Exception:
        return None

    # --------------------------------------------------------
    # Training statistics
    # --------------------------------------------------------

    fitted_values = ols_model.predict(x_sm)
    residuals = y_clean - fitted_values

    training_rmse = np.sqrt(
        mean_squared_error(y_clean, fitted_values)
    )

    # --------------------------------------------------------
    # Leave-one-out cross-validation
    # --------------------------------------------------------

    try:
        loo = LeaveOneOut()
        cv_model = LinearRegression()

        loocv_predictions = cross_val_predict(
            cv_model,
            x_poly,
            y_clean,
            cv=loo
        )

        loocv_rmse = np.sqrt(
            mean_squared_error(
                y_clean,
                loocv_predictions
            )
        )

        ss_res_cv = np.sum(
            (y_clean - loocv_predictions) ** 2
        )

        ss_tot_cv = np.sum(
            (y_clean - np.mean(y_clean)) ** 2
        )

        loocv_r2 = (
            1 - ss_res_cv / ss_tot_cv
            if ss_tot_cv > 0
            else np.nan
        )

    except Exception:
        loocv_predictions = np.full(n, np.nan)
        loocv_rmse = np.nan
        loocv_r2 = np.nan

    # --------------------------------------------------------
    # Residual diagnostics
    # --------------------------------------------------------

    shapiro_stat, shapiro_p = safe_shapiro(residuals)

    (
        bp_lm_stat,
        bp_lm_p,
        bp_f_stat,
        bp_f_p
    ) = safe_breusch_pagan(
        residuals,
        x_sm
    )

    diagnostics_status = diagnostic_status(
        shapiro_p,
        bp_lm_p
    )

    # --------------------------------------------------------
    # ANOVA quantities
    # --------------------------------------------------------

    total_ss = np.sum(
        (y_clean - np.mean(y_clean)) ** 2
    )

    residual_ss = np.sum(residuals ** 2)

    regression_ss = total_ss - residual_ss

    df_model = int(ols_model.df_model)
    df_residual = int(ols_model.df_resid)
    df_total = n - 1

    ms_regression = (
        regression_ss / df_model
        if df_model > 0
        else np.nan
    )

    ms_residual = (
        residual_ss / df_residual
        if df_residual > 0
        else np.nan
    )

    residual_standard_error = (
        np.sqrt(ms_residual)
        if not pd.isna(ms_residual)
        else np.nan
    )

    # --------------------------------------------------------
    # Identifiers
    # --------------------------------------------------------

    model_type = model_type_from_degree(degree)

    model_id = (
        f"{property_name} | "
        f"{index_name} | "
        f"{model_type}"
    )

    equation = make_equation(
        property_name,
        ols_model.params,
        index_name
    )

    # --------------------------------------------------------
    # Main summary row
    # --------------------------------------------------------

    summary_row = {
        "Model ID": model_id,
        "Property": property_name,
        "Topological Index": index_name,
        "Model Type": model_type,
        "Degree": degree,
        "Equation": equation,
        "n": n,
        "df Model": df_model,
        "df Residual": df_residual,
        "R2": ols_model.rsquared,
        "Adjusted R2": ols_model.rsquared_adj,
        "Training RMSE": training_rmse,
        "LOOCV RMSE": loocv_rmse,
        "LOOCV R2": loocv_r2,
        "AIC": ols_model.aic,
        "BIC": ols_model.bic,
        "F-statistic": ols_model.fvalue,
        "Model p-value": ols_model.f_pvalue,
        "Residual SE": residual_standard_error,
        "Shapiro p-value": shapiro_p,
        "Breusch-Pagan p-value": bp_lm_p,
        "Diagnostic Status": diagnostics_status
    }

    # --------------------------------------------------------
    # Coefficients table
    # --------------------------------------------------------

    coefficient_rows = []

    term_names = ["Intercept"]

    for power in range(1, degree + 1):
        if power == 1:
            term_names.append("X")
        else:
            term_names.append(f"X^{power}")

    confidence_intervals = ols_model.conf_int()

    for position, term in enumerate(term_names):
        coefficient_rows.append({
            "Model ID": model_id,
            "Property": property_name,
            "Topological Index": index_name,
            "Model Type": model_type,
            "Term": term,
            "Coefficient": ols_model.params[position],
            "Std. Error": ols_model.bse[position],
            "t-statistic": ols_model.tvalues[position],
            "p-value": ols_model.pvalues[position],
            "95% CI Lower": confidence_intervals[position, 0],
            "95% CI Upper": confidence_intervals[position, 1]
        })

    # --------------------------------------------------------
    # ANOVA table
    # --------------------------------------------------------

    anova_rows = [
        {
            "Model ID": model_id,
            "Property": property_name,
            "Topological Index": index_name,
            "Model Type": model_type,
            "Source": "Regression",
            "Sum of Squares": regression_ss,
            "df": df_model,
            "Mean Square": ms_regression,
            "F-statistic": ols_model.fvalue,
            "p-value": ols_model.f_pvalue
        },
        {
            "Model ID": model_id,
            "Property": property_name,
            "Topological Index": index_name,
            "Model Type": model_type,
            "Source": "Residual",
            "Sum of Squares": residual_ss,
            "df": df_residual,
            "Mean Square": ms_residual,
            "F-statistic": np.nan,
            "p-value": np.nan
        },
        {
            "Model ID": model_id,
            "Property": property_name,
            "Topological Index": index_name,
            "Model Type": model_type,
            "Source": "Total",
            "Sum of Squares": total_ss,
            "df": df_total,
            "Mean Square": np.nan,
            "F-statistic": np.nan,
            "p-value": np.nan
        }
    ]

    # --------------------------------------------------------
    # VIF table
    # --------------------------------------------------------

    vif_rows = []

    for term, vif_value in calculate_vif(x_sm, degree):
        vif_rows.append({
            "Model ID": model_id,
            "Property": property_name,
            "Topological Index": index_name,
            "Model Type": model_type,
            "Term": term,
            "VIF": vif_value
        })

    # --------------------------------------------------------
    # Diagnostics table
    # --------------------------------------------------------

    diagnostics_row = {
        "Model ID": model_id,
        "Property": property_name,
        "Topological Index": index_name,
        "Model Type": model_type,
        "Shapiro-Wilk Statistic": shapiro_stat,
        "Shapiro p-value": shapiro_p,
        "Breusch-Pagan LM Statistic": bp_lm_stat,
        "Breusch-Pagan p-value": bp_lm_p,
        "Breusch-Pagan F Statistic": bp_f_stat,
        "Breusch-Pagan F p-value": bp_f_p,
        "Mean Residual": np.mean(residuals),
        "Residual SD": np.std(residuals, ddof=1),
        "Diagnostic Status": diagnostics_status
    }

    # --------------------------------------------------------
    # Observed and predicted values
    # --------------------------------------------------------

    prediction_rows = []

    for (
        compound,
        observed,
        training_predicted,
        training_residual,
        loocv_predicted
    ) in zip(
        compounds_clean,
        y_clean,
        fitted_values,
        residuals,
        loocv_predictions
    ):

        loocv_residual = (
            observed - loocv_predicted
            if not pd.isna(loocv_predicted)
            else np.nan
        )

        prediction_rows.append({
            "Model ID": model_id,
            "Property": property_name,
            "Topological Index": index_name,
            "Model Type": model_type,
            "Compound": compound,
            "Observed": observed,
            "Training Predicted": training_predicted,
            "Training Residual": training_residual,
            "LOOCV Predicted": loocv_predicted,
            "LOOCV Residual": loocv_residual
        })

    return {
        "summary": summary_row,
        "coefficients": coefficient_rows,
        "anova": anova_rows,
        "vif": vif_rows,
        "diagnostics": diagnostics_row,
        "predictions": prediction_rows
    }


# ============================================================
# 4. MODEL SELECTION FUNCTIONS
# ============================================================

def select_recommended_models(model_summary_df):
    """
    Final model-selection rule:

    1. Keep models with significant overall OLS F-test, p < 0.05.
    2. Select the lowest BIC model for each property.
    3. If BIC is equal, use lower LOOCV RMSE.
    4. If still tied, use higher LOOCV R2.
    5. If still tied, use the simpler polynomial degree.
    """

    selected_rows = []

    for property_name in PROPERTIES:
        property_models = model_summary_df[
            model_summary_df["Property"] == property_name
        ].copy()

        significant_models = property_models[
            property_models["Model p-value"] < ALPHA
        ].copy()

        if significant_models.empty:
            candidates = property_models.copy()

            selection_rule = (
                "No significant model found; selected lowest BIC "
                "among all fitted models"
            )
        else:
            candidates = significant_models.copy()

            selection_rule = (
                "Lowest BIC among models with overall F-test p<0.05"
            )

        candidates = candidates.sort_values(
            by=[
                "BIC",
                "LOOCV RMSE",
                "LOOCV R2",
                "Degree"
            ],
            ascending=[
                True,
                True,
                False,
                True
            ],
            na_position="last"
        )

        selected = candidates.iloc[0].copy()

        selected["Selection Rule"] = selection_rule

        if pd.isna(selected["LOOCV R2"]):
            selected["Predictive Reliability"] = "Not assessed"

        elif selected["LOOCV R2"] > LOOCV_R2_THRESHOLD:
            selected["Predictive Reliability"] = "Reliable"

        else:
            selected["Predictive Reliability"] = (
                "Not reliable: LOOCV R2 <= 0"
            )

        selected_rows.append(selected)

    recommended_df = pd.DataFrame(selected_rows)

    order_map = {
        property_name: number
        for number, property_name in enumerate(PROPERTIES)
    }

    recommended_df["Property Order"] = (
        recommended_df["Property"].map(order_map)
    )

    recommended_df = (
        recommended_df
        .sort_values("Property Order")
        .drop(columns="Property Order")
        .reset_index(drop=True)
    )

    return recommended_df


def select_best_by_metric(
    model_summary_df,
    metric,
    ascending=True,
    significant_only=False
):
    """Select one best model per property using a requested metric."""

    selected_rows = []

    for property_name in PROPERTIES:
        candidates = model_summary_df[
            model_summary_df["Property"] == property_name
        ].copy()

        if significant_only:
            significant = candidates[
                candidates["Model p-value"] < ALPHA
            ].copy()

            if not significant.empty:
                candidates = significant

        candidates = candidates.sort_values(
            by=metric,
            ascending=ascending,
            na_position="last"
        )

        selected_rows.append(candidates.iloc[0])

    selected_df = pd.DataFrame(selected_rows)

    order_map = {
        property_name: number
        for number, property_name in enumerate(PROPERTIES)
    }

    selected_df["Property Order"] = (
        selected_df["Property"].map(order_map)
    )

    selected_df = (
        selected_df
        .sort_values("Property Order")
        .drop(columns="Property Order")
        .reset_index(drop=True)
    )

    return selected_df


# ============================================================
# 5. EXCEL FORMATTING
# ============================================================

def format_workbook(file_path):
    """Apply formatting using openpyxl after all sheets are written."""

    workbook = load_workbook(file_path)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7"
    )

    header_font = Font(
        bold=True,
        color="000000"
    )

    thin_side = Side(
        style="thin",
        color="B7B7B7"
    )

    border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = False

        # Header row
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )
            cell.border = border

        # Body cells
        for row in worksheet.iter_rows(
            min_row=2,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=worksheet.max_column
        ):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=False
                )

                if isinstance(cell.value, (int, float, np.number)):
                    cell.number_format = "0.0000"

        # Enable Excel filter
        worksheet.auto_filter.ref = worksheet.dimensions

        # Adjust column widths
        for column_cells in worksheet.columns:
            column_letter = get_column_letter(
                column_cells[0].column
            )

            header = str(column_cells[0].value)

            maximum_length = len(header)

            for cell in column_cells[1:]:
                if cell.value is not None:
                    maximum_length = max(
                        maximum_length,
                        len(str(cell.value))
                    )

            width = min(max(maximum_length + 2, 12), 35)

            if header == "Equation":
                width = 80

            elif header == "Model ID":
                width = 32

            elif header == "Selection Rule":
                width = 58

            elif header in [
                "Topological Index",
                "Predictive Reliability",
                "Diagnostic Status"
            ]:
                width = 30

            worksheet.column_dimensions[column_letter].width = width

    workbook.save(file_path)


# ============================================================
# 6. MAIN PROGRAM
# ============================================================

def main():

    print("=" * 70)
    print("FULL OLS QSPR ANALYSIS")
    print("=" * 70)

    print(f"Input file: {INPUT_FILE}")
    print(f"Output workbook: {OUTPUT_FILE}")

    # Ensure project folder exists
    PROJECT_FOLDER.mkdir(parents=True, exist_ok=True)

    # Confirm input data file exists
    if not INPUT_FILE.exists():
        raise FileNotFoundError(
            "\nInput Excel file was not found.\n\n"
            f"Expected location:\n{INPUT_FILE}\n\n"
            "Check INPUT_FILE_NAME near the top of this script."
        )

    # --------------------------------------------------------
    # Read dataset
    # --------------------------------------------------------

    df = pd.read_excel(INPUT_FILE)

    if COMPOUND_COLUMN not in df.columns:
        raise ValueError(
            "\nCompound column was not found.\n\n"
            f"Expected: {COMPOUND_COLUMN}\n\n"
            f"Available columns:\n{list(df.columns)}"
        )

    missing_properties = [
        property_name
        for property_name in PROPERTIES
        if property_name not in df.columns
    ]

    if missing_properties:
        raise ValueError(
            "\nThe following required property columns are missing:\n"
            f"{missing_properties}"
        )

    compound_names = (
        df[COMPOUND_COLUMN]
        .astype(str)
        .to_numpy()
    )

    # Identify numerical descriptor columns
    index_columns = [
        column
        for column in df.columns
        if (
            column not in PROPERTIES
            and column != COMPOUND_COLUMN
            and pd.api.types.is_numeric_dtype(df[column])
        )
    ]

    if len(index_columns) == 0:
        raise ValueError(
            "No numerical topological-index columns were detected."
        )

    expected_model_count = (
        len(PROPERTIES)
        * len(index_columns)
        * len(DEGREES)
    )

    print(f"Compound-name column: {COMPOUND_COLUMN}")
    print(f"Number of compounds: {len(df)}")
    print(f"Number of properties: {len(PROPERTIES)}")
    print(f"Number of topological indices: {len(index_columns)}")
    print(f"Expected model count: {expected_model_count}")

    # --------------------------------------------------------
    # Pearson correlation analysis
    # --------------------------------------------------------

    correlation_rows = []

    for property_name in PROPERTIES:
        y = df[property_name].to_numpy(dtype=float)

        for index_name in index_columns:
            x = df[index_name].to_numpy(dtype=float)

            valid_mask = np.isfinite(x) & np.isfinite(y)

            x_clean = x[valid_mask]
            y_clean = y[valid_mask]

            if len(x_clean) > 2 and np.std(x_clean) > 0:
                r_value, p_value = pearsonr(
                    x_clean,
                    y_clean
                )

                correlation_rows.append({
                    "Property": property_name,
                    "Topological Index": index_name,
                    "Pearson r": r_value,
                    "Absolute r": abs(r_value),
                    "p-value": p_value,
                    "n": len(x_clean)
                })

    correlation_df = pd.DataFrame(correlation_rows)

    correlation_df = (
        correlation_df
        .sort_values(
            by=["Property", "Absolute r"],
            ascending=[True, False]
        )
        .reset_index(drop=True)
    )

    top_correlations_df = (
        correlation_df
        .groupby(
            "Property",
            group_keys=False
        )
        .head(TOP_CORRELATIONS_TO_KEEP)
        .reset_index(drop=True)
    )

    # --------------------------------------------------------
    # Correlation matrix
    # --------------------------------------------------------

    correlation_matrix_df = pd.DataFrame(
        index=index_columns,
        columns=PROPERTIES,
        dtype=float
    )

    for property_name in PROPERTIES:
        property_correlations = correlation_df[
            correlation_df["Property"] == property_name
        ].set_index("Topological Index")

        correlation_matrix_df[property_name] = (
            correlation_matrix_df.index.map(
                property_correlations["Pearson r"]
            )
        )

    correlation_matrix_df = (
        correlation_matrix_df
        .reset_index()
        .rename(columns={"index": "Topological Index"})
    )

    # --------------------------------------------------------
    # RMS of topological indices
    # --------------------------------------------------------

    rms_rows = []

    for index_name in index_columns:
        values = df[index_name].to_numpy(dtype=float)
        values = values[np.isfinite(values)]

        rms_value = (
            np.sqrt(np.mean(values ** 2))
            if len(values) > 0
            else np.nan
        )

        rms_rows.append({
            "Topological Index": index_name,
            "RMS": rms_value
        })

    rms_df = (
        pd.DataFrame(rms_rows)
        .sort_values("RMS", ascending=False)
        .reset_index(drop=True)
    )

    # --------------------------------------------------------
    # Fit all models
    # --------------------------------------------------------

    all_model_rows = []
    all_coefficient_rows = []
    all_anova_rows = []
    all_vif_rows = []
    all_diagnostic_rows = []
    all_prediction_rows = []

    for property_name in PROPERTIES:
        print(f"Fitting {property_name} models...")

        y = df[property_name].to_numpy(dtype=float)

        for index_name in index_columns:
            x = df[index_name].to_numpy(dtype=float)

            for degree in DEGREES:
                result = fit_ols_polynomial_model(
                    x=x,
                    y=y,
                    property_name=property_name,
                    index_name=index_name,
                    compounds=compound_names,
                    degree=degree
                )

                if result is None:
                    continue

                all_model_rows.append(result["summary"])
                all_coefficient_rows.extend(
                    result["coefficients"]
                )
                all_anova_rows.extend(result["anova"])
                all_vif_rows.extend(result["vif"])
                all_diagnostic_rows.append(
                    result["diagnostics"]
                )
                all_prediction_rows.extend(
                    result["predictions"]
                )

    model_summary_df = pd.DataFrame(all_model_rows)
    coefficient_tests_df = pd.DataFrame(all_coefficient_rows)
    anova_df = pd.DataFrame(all_anova_rows)
    vif_df = pd.DataFrame(all_vif_rows)
    diagnostics_df = pd.DataFrame(all_diagnostic_rows)
    predictions_df = pd.DataFrame(all_prediction_rows)

    print(f"Number of fitted models: {len(model_summary_df)}")

    # --------------------------------------------------------
    # Create model-selection sheets
    # --------------------------------------------------------

    significant_models_df = (
        model_summary_df[
            model_summary_df["Model p-value"] < ALPHA
        ]
        .copy()
        .sort_values(
            by=[
                "Property",
                "BIC",
                "LOOCV RMSE",
                "Degree"
            ],
            ascending=[
                True,
                True,
                True,
                True
            ],
            na_position="last"
        )
        .reset_index(drop=True)
    )

    diagnostic_pass_models_df = (
        significant_models_df[
            significant_models_df["Diagnostic Status"] == "Pass"
        ]
        .copy()
        .reset_index(drop=True)
    )

    ranked_candidates_df = (
        significant_models_df
        .sort_values(
            by=[
                "Property",
                "BIC",
                "LOOCV RMSE",
                "LOOCV R2",
                "Degree"
            ],
            ascending=[
                True,
                True,
                True,
                False,
                True
            ],
            na_position="last"
        )
        .reset_index(drop=True)
    )

    recommended_models_df = select_recommended_models(
        model_summary_df
    )

    best_by_bic_df = select_best_by_metric(
        model_summary_df,
        metric="BIC",
        ascending=True,
        significant_only=True
    )

    best_by_aic_df = select_best_by_metric(
        model_summary_df,
        metric="AIC",
        ascending=True,
        significant_only=True
    )

    best_by_loocv_df = select_best_by_metric(
        model_summary_df,
        metric="LOOCV RMSE",
        ascending=True,
        significant_only=False
    )

    best_by_adjusted_r2_df = select_best_by_metric(
        model_summary_df,
        metric="Adjusted R2",
        ascending=False,
        significant_only=False
    )

    # --------------------------------------------------------
    # Extract supporting output for final recommended models
    # --------------------------------------------------------

    recommended_model_ids = set(
        recommended_models_df["Model ID"]
    )

    recommended_coefficients_df = (
        coefficient_tests_df[
            coefficient_tests_df["Model ID"]
            .isin(recommended_model_ids)
        ]
        .copy()
    )

    recommended_anova_df = (
        anova_df[
            anova_df["Model ID"]
            .isin(recommended_model_ids)
        ]
        .copy()
    )

    recommended_vif_df = (
        vif_df[
            vif_df["Model ID"]
            .isin(recommended_model_ids)
        ]
        .copy()
    )

    recommended_diagnostics_df = (
        diagnostics_df[
            diagnostics_df["Model ID"]
            .isin(recommended_model_ids)
        ]
        .copy()
    )

    recommended_predictions_df = (
        predictions_df[
            predictions_df["Model ID"]
            .isin(recommended_model_ids)
        ]
        .copy()
    )

    # --------------------------------------------------------
    # Summary sheet
    # --------------------------------------------------------

    summary_df = pd.DataFrame({
        "Item": [
            "Analysis type",
            "Regression method",
            "Input data file",
            "Output workbook",
            "Number of compounds",
            "Number of physicochemical properties",
            "Number of topological indices",
            "Polynomial forms fitted",
            "Expected number of models",
            "Number of fitted models",
            "Overall F-test significance threshold",
            "Final selection rule",
            "Predictive reliability criterion"
        ],
        "Details": [
            "Quantitative structure-property relationship analysis",
            "Ordinary least squares regression",
            str(INPUT_FILE),
            str(OUTPUT_FILE),
            len(df),
            len(PROPERTIES),
            len(index_columns),
            "Linear, quadratic, and cubic",
            expected_model_count,
            len(model_summary_df),
            f"p < {ALPHA}",
            (
                "Lowest BIC among models with significant "
                "overall F-test; LOOCV RMSE, LOOCV R2, and "
                "degree used as tie-breakers"
            ),
            (
                "LOOCV R2 > 0 indicates positive cross-validated "
                "predictive performance"
            )
        ]
    })

    # --------------------------------------------------------
    # Write Excel workbook
    # --------------------------------------------------------

    print("\nWriting workbook. Please do not close Python.")

    try:
        with pd.ExcelWriter(
            OUTPUT_FILE,
            engine="openpyxl",
            mode="w"
        ) as writer:

            summary_df.to_excel(
                writer,
                sheet_name="Summary",
                index=False
            )

            recommended_models_df.to_excel(
                writer,
                sheet_name="Recommended Models",
                index=False
            )

            best_by_bic_df.to_excel(
                writer,
                sheet_name="Best by BIC",
                index=False
            )

            best_by_aic_df.to_excel(
                writer,
                sheet_name="Best by AIC",
                index=False
            )

            best_by_loocv_df.to_excel(
                writer,
                sheet_name="Best by LOOCV",
                index=False
            )

            best_by_adjusted_r2_df.to_excel(
                writer,
                sheet_name="Best by Adj R2",
                index=False
            )

            correlation_df.to_excel(
                writer,
                sheet_name="Correlation Results",
                index=False
            )

            top_correlations_df.to_excel(
                writer,
                sheet_name="Top Correlations",
                index=False
            )

            correlation_matrix_df.to_excel(
                writer,
                sheet_name="Correlation Matrix",
                index=False
            )

            rms_df.to_excel(
                writer,
                sheet_name="RMS Indices",
                index=False
            )

            model_summary_df.to_excel(
                writer,
                sheet_name="All Models Summary",
                index=False
            )

            significant_models_df.to_excel(
                writer,
                sheet_name="Significant Models",
                index=False
            )

            diagnostic_pass_models_df.to_excel(
                writer,
                sheet_name="Diagnostic Pass Models",
                index=False
            )

            ranked_candidates_df.to_excel(
                writer,
                sheet_name="Ranked Candidates",
                index=False
            )

            coefficient_tests_df.to_excel(
                writer,
                sheet_name="Coefficient Tests",
                index=False
            )

            anova_df.to_excel(
                writer,
                sheet_name="ANOVA",
                index=False
            )

            vif_df.to_excel(
                writer,
                sheet_name="VIF",
                index=False
            )

            diagnostics_df.to_excel(
                writer,
                sheet_name="Residual Diagnostics",
                index=False
            )

            predictions_df.to_excel(
                writer,
                sheet_name="Predictions",
                index=False
            )

            recommended_coefficients_df.to_excel(
                writer,
                sheet_name="Recommended Coefficients",
                index=False
            )

            recommended_anova_df.to_excel(
                writer,
                sheet_name="Recommended ANOVA",
                index=False
            )

            recommended_vif_df.to_excel(
                writer,
                sheet_name="Recommended VIF",
                index=False
            )

            recommended_diagnostics_df.to_excel(
                writer,
                sheet_name="Recommended Diagnostics",
                index=False
            )

            recommended_predictions_df.to_excel(
                writer,
                sheet_name="Recommended Predictions",
                index=False
            )

    except Exception as excel_error:
        raise RuntimeError(
            "\nExcel workbook could not be created.\n\n"
            f"Reason: {excel_error}"
        )

    # Apply workbook formatting
    print("Applying worksheet formatting...")
    format_workbook(OUTPUT_FILE)

    # Confirm the workbook truly exists
    if not OUTPUT_FILE.exists():
        raise FileNotFoundError(
            "\nWorkbook was not found after export.\n\n"
            f"Expected location:\n{OUTPUT_FILE}"
        )

    print("\n" + "=" * 70)
    print("ANALYSIS COMPLETED SUCCESSFULLY")
    print("=" * 70)
    print(f"Workbook created at:\n{OUTPUT_FILE}")

    print("\nRecommended models:")

    columns_to_show = [
        "Property",
        "Topological Index",
        "Model Type",
        "Training RMSE",
        "LOOCV RMSE",
        "LOOCV R2",
        "Adjusted R2",
        "BIC",
        "Model p-value",
        "Predictive Reliability"
    ]

    print(
        recommended_models_df[
            columns_to_show
        ].to_string(index=False)
    )


# ============================================================
# 7. RUN THE PROGRAM
# ============================================================

if __name__ == "__main__":
    try:
        main()

    except Exception as error:
        print("\n" + "=" * 70)
        print("ERROR")
        print("=" * 70)
        print(error)

        sys.exit(1)